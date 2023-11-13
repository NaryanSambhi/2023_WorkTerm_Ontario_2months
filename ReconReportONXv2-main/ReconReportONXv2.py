# Naryan Sambhi - 6/29/2023
#
# goal:
# create a new version of reconreportonx that must:
#
#   - Get files from email 
#   - Merge files into a single excel document & create report 
#   - Send a formated email with attatched docs

#header
import win32com.client as client
import datetime
import os
import time
import openpyxl
from copy import copy
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from dateutil import parser
import ctypes
import pythoncom
from openpyxl.styles import Font
from openpyxl.styles import Alignment 
from openpyxl.utils import get_column_letter

# main code body found at bottom (compiles all functions into a body)
# each area is like its own file

###################################################### FILE PATHS - CONFIGURE TO YOUR MACHINE ######################################################

#user email and targeted folder
UserEmail = 'Naryan.Sambhi@ontario.ca'
UserFolder = 'Recon'

#path to where files will save
basefilepath = rf'C:\Users\SambhiNa\OneDrive - Government of Ontario\Desktop\Otest_path'

#file path of used files
ReconWorkOrderFP = rf'C:\Users\SambhiNa\OneDrive - Government of Ontario\Desktop\Otest_path\Recon Report WO.xlsx'
ReconReportFP = rf'C:\Users\SambhiNa\OneDrive - Government of Ontario\Desktop\Otest_path\Recon Report.xlsx'

#where final report will save to & as 
ReconReportONX = rf'C:\Users\SambhiNa\OneDrive - Government of Ontario\Desktop\Otest_path\ReconReportONX.xlsx'


#subject email key 1
subject1 = 'eSMT Reconciliation Report'  #will also find eSMT Reconciliation Report WO


#NEW REPORT email file recepients 
messageTO = 'naryan.sambhi@ontario.ca'
messageCC = ''
messageSUBJECT = 'Reconciliation Report ONX'



#how long system waits 
sleeptime = 1 

#num of days difference before raising an error
days = 0 #0 day difference befor errors are raised 


###################################################### GENERAL FUNCTIONS ######################################################

#check if file exists 
def Exists(file):
    try:
        ReconReport = time.ctime(os.path.getmtime(file))
        #print(rf"File x has been found successfully with date: {ReconReport}")  
    except(WindowsError):
        print("\n*************************************************************************************************")
        print("\nERROR!\nFILE DOES NOT EXIST:\n%s\n\n" %(file))
        print("*************************************************************************************************")
        print("\nExiting Program...\n")
        exit(1)



#get difference from current day and compare to email day, if more than desired day, exit
def difdatesOutlook(message): #used in get files

        dateMessage = message.SentOn
    
        #get the current date 
        dateNow = datetime.datetime.now()

        #remove time zone info of email date
        dateMessage = dateMessage.replace(tzinfo=None)

        #remove the hours, minutes and seconds to calculate de difference of days just by day
        dateMessageRound = dateMessage.replace(hour=0, minute=0, second=0, microsecond=0)
        dateNowRound = dateNow.replace(hour=0, minute=0, second=0, microsecond=0)
       
        #get the dif of day between actual message in the loop and the current date
        difDay = dateNowRound - dateMessageRound

        #check, if more than x days of difference, exit with error
        if difDay > datetime.timedelta(days):
            print("ERROR!\nMOST RECENT FILES ARE OUTDATED\n\n*************************************************************************************************")
            print("\nExiting Program...\n")
            exit(2)




#get difference from current day and compare to email day, if more than desired day, exit
def difdatesFiles(file):
    
    #get current day
    dateNow = datetime.datetime.now()
    dateValidateFile = dateNow.strftime("%a %b %d %H:%M:%S %Y")

    #get date of file to validate
    File = time.ctime(os.path.getmtime(file))

    #creation of workable variables 
    dateNowTovalidate = parser.parse(dateValidateFile)
    dateFile = parser.parse(File)

    #delete hours and minutes from file and current date
    rounded_Now = dateNowTovalidate.replace(hour=0, minute=0, second=0, microsecond=0)
    rounded_File = dateFile.replace(hour=0, minute=0, second=0, microsecond=0)

    #diff in days
    diff = rounded_Now - rounded_File

    #check, if more than x days of difference, exit with error
    if  diff > datetime.timedelta(days):
        print("\n*************************************************************************************************")
        print("\nERROR!\nFILE IS OUTDATED:\n%s\n\n" %(file))
        print("*************************************************************************************************")
        print("\nExiting Program...\n")
        exit(2)




#get the user info from windows - function taken directly from ONXv1
def get_data(EXTENDED_NAME_FORMAT: int):
    GetUserNameEx = ctypes.windll.secur32.GetUserNameExW
    data = EXTENDED_NAME_FORMAT

    size = ctypes.pointer(ctypes.c_ulong(0))
    GetUserNameEx(data, None, size)

    nameBuffer = ctypes.create_unicode_buffer(size.contents.value)
    GetUserNameEx(data, nameBuffer, size)
    return nameBuffer.value



###################################################### ONX SPECIFIC FUNCTIONS ######################################################

# Enter Outlook and Recieves Files, 
# Arguments: 
# number of emails found until stopped 
# name of subject to find 
#email = UserEmail 
#folder = UserFolder

def getFiles(number_of_emails, email_name, user_email, user_folder): #email name is subject in code execution 


    #connect to outlook
    print("\nGetting files from outlook..")
    print('\n************************************************************************************************\n')

    #get connction with outlook
    outlook = client.Dispatch("Outlook.Application", pythoncom.CoInitialize()).GetNamespace("MAPI")
    outlook.SendAndReceive

    #get account and folders in Outlook
    account = outlook.Folders[user_email]
    inboxNodedown = account.Folders[user_folder]
    messages = inboxNodedown.Items


    #checks if folder is empty or not
    #important after weekends -> files can be put into cloud automatically, and box could be empty causing errors in code
    if len(messages) == 0:
        print("ERROR!\n\nNO MESSAGES DETECTED IN INBOX:\n\nEMAIL: %s \nFOLDER: %s" %(UserEmail, UserFolder))
        print("\n\n************************************************************************************************\n")
        exit(1)

    #Sort the messages by receive time
    messages.Sort("[ReceivedTime]", True)

    #get "Recon Report.xlsx & WO"
    count = 0 #number of times looped 
    for message in messages:
        subject = message.Subject         
        attachments = message.Attachments

        difdatesOutlook(message) #checking difference in dates and exiting if issues arise 

        if email_name in subject:      
            count += 1
            print("Email %d, found" %(count))

            for attachment in attachments:
                pathFile = os.path.join(basefilepath, str(attachment))
                attachment.SaveAsFile(pathFile)
                print(rf"Name file Saved: {pathFile}")      
                #print("File Saved: %s" %(pathFile))

            #check if file exists (stop garbage in)
            Exists(pathFile) #will exit program if file is not found 

            print('\n************************************************************************************************\n')

            time.sleep(sleeptime)   

            #stop once NUM of emails found finished
            if count == number_of_emails:
                print("\nAll Files Found Successfully\n")
                print('\n************************************************************************************************\n')
                time.sleep(sleeptime)
                break    
        else:
            print("Skipped file")





#merge files into a single file for report - based off of legacy code
def mergeFiles(file1, file2, save_path): #file 1 #file 2 #xlsx

#check files
    #double checking if files are found correctly or raise an error and exist
    Exists(file1)
    Exists(file2)

    time.sleep(sleeptime)

    #checking if files are of current date
    difdatesFiles(file1)
    difdatesFiles(file2)
    
    #ui
    print("\nmerging files...\n\n")
    print("*************************************************************************************************\n")
    time.sleep(sleeptime)


    wbReport = openpyxl.Workbook()

    #create spreadsheets
    sheetInc = wbReport.create_sheet("INC")
    sheetWO = wbReport.create_sheet("WO")


    #load the file with incidents
    wbInc = openpyxl.load_workbook(file1)
    sheetIncSource = wbInc.active
        
    #load the file with work orders
    wbWO = openpyxl.load_workbook(file2)

    sheetWOSource = wbWO.active

    indexRow = 0



############################################ COPY PASTE FROM RECONREPORT.PY AND ONX V1 #################################

    #delete the rows that does not contain ONX INCIDENT SUPPORT
    for row in range(2, sheetIncSource.max_row + 1):
        
        #verify if there a row deleted
        if indexRow > 0:
            #set the new row according to the number of rows previously deleted
            row = row - indexRow
        assignedGroup = sheetIncSource.cell(row, 7).value

                
        if assignedGroup != "ONX INCIDENT SUPPORT":              
            sheetIncSource.delete_rows(row, 1)
            indexRow = indexRow + 1
            

    #create a cell style for even cells
    highlightEven = NamedStyle(name="highlightEven")
    highlightEven.font = Font(name='Arial', bold=False, size=10)
    highlightEven.fill = PatternFill(start_color="ECEAEC", end_color="ECEAEC", fill_type = "solid")

    #create a cell style for odd cells
    highlightOdd = NamedStyle(name="highlightOdd")
    highlightOdd.font = Font(name='Arial', bold=False, size=10)
    highlightOdd.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")

    #create a cell style for hypelinks odd cells
    incHyperlink = NamedStyle(name="incHyperlink")
    incHyperlink.font = Font(name='Arial', bold=False, size=10, color='0000FF', underline='single')

    #create a cell style for hypelinks even cells
    incHyperlinkEven = NamedStyle(name="incHyperlinkEven")
    incHyperlinkEven.font = Font(name='Arial', bold=False, size=10, color='0000FF', underline='single')
    incHyperlinkEven.fill = PatternFill(start_color="ECEAEC", end_color="ECEAEC", fill_type = "solid")


    #Initialize header row, for the INC sheet Target Date
    for col in range(1, 12):
        #set the names for the header
        sheetInc.cell(1, col).value = sheetIncSource.cell(1, col).value 

        #fix the name in column 9
        if col == 9:
            sheetInc.cell(1, col).value = "Target Date"          
        
        #set th styles for header
        sheetInc.cell(1, col).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
        sheetInc.cell(1, col).border = copy(sheetIncSource.cell(1, col).border)
        sheetInc.cell(1, col).fill = copy(sheetIncSource.cell(1, col).fill)
        sheetInc.cell(1, col).number_format = copy(sheetIncSource.cell(1, col).number_format)
        sheetInc.cell(1, col).alignment = copy(sheetIncSource.cell(1, col).alignment)

    #add a new column = Assignment Group for ONX
    sheetInc.cell(1, 11).value = "Assignment Group ONX" 
    sheetInc.cell(1, 11).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 11).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 11).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 11).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 11).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")

        #add a new column = State for ONX
    sheetInc.cell(1, 12).value = "State ONX" 
    sheetInc.cell(1, 12).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 12).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 12).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 12).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 12).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")

        #add a new column = Comments for ONX
    sheetInc.cell(1, 13).value = "Comments ONX" 
    sheetInc.cell(1, 13).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 13).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 13).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 13).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 13).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")

    #add a new column = Last Modified Date for ONX
    sheetInc.cell(1, 14).value = "Last Modified Date ONX" 
    sheetInc.cell(1, 14).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 14).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 14).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 14).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 14).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")

        #add a new column = Priority for ONX
    sheetInc.cell(1, 15).value = "Priority ONX" 
    sheetInc.cell(1, 15).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 15).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 15).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 15).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 15).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")

    #add a new column = OCC Commemt
    sheetInc.cell(1, 16).value = "OCC Comments" 
    sheetInc.cell(1, 16).border = copy(sheetIncSource.cell(1, 10).border)
    sheetInc.cell(1, 16).fill = copy(sheetIncSource.cell(1, 10).fill)
    sheetInc.cell(1, 16).number_format = copy(sheetIncSource.cell(1, 10).number_format)
    sheetInc.cell(1, 16).alignment = copy(sheetIncSource.cell(1, 10).alignment)
    sheetInc.cell(1, 16).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")



    indexRow = 0

    #loop for to check each rows and col of file
    for row in range(2, sheetIncSource.max_row + 1):
        for col in range(1, 17):
            
            assignedGroup = sheetIncSource.cell(row, 7).value                    

            if assignedGroup == "ONX INCIDENT SUPPORT":
                
                if indexRow > 0:
                    row = row + indexRow

                #set the value in the cell
                sheetInc.cell(row, col).value = sheetIncSource.cell(row, col).value

                #verify if the cell has hyperlink and set styles
                if sheetIncSource.cell(row, col).hyperlink:
                    sheetInc.cell(row, col)._hyperlink = copy(sheetIncSource.cell(row, col).hyperlink)
                    sheetInc.cell(row, col).style = incHyperlink
                #verify if the row is even 
                if row % 2 == 0:
                    sheetInc.cell(row, col).style = highlightEven
                    if col == 1:
                        sheetInc.cell(row, col).style = incHyperlinkEven
                else:
                    sheetInc.cell(row, col).style = highlightOdd

                    if col == 1:
                        sheetInc.cell(row, col).style = incHyperlink
                #set the format to clls with dates and stye
                if col == 2 or col == 3:
                    sheetInc.cell(row, col).number_format = copy(sheetIncSource.cell(row, col).number_format)
                    sheetInc.cell(row, col).alignment = copy(sheetIncSource.cell(row, col).alignment)

                #get the values of submitdate and modifiedDate
                submitDate = sheetIncSource.cell(row, 2).value
                lastModifiedDate = sheetIncSource.cell(row, 3).value

                #verify that the dates are not none
                if submitDate is not None and lastModifiedDate is not None:
                
                    #delete the hour of dates to calculate the difference of days
                    submitDate = submitDate.replace(hour=0, minute=0, second=0, microsecond=0)
                    lastModifiedDate = lastModifiedDate.replace(hour=0, minute=0, second=0, microsecond=0)
                    dateNowReport = datetime.datetime.now()
                    dateNowReport = dateNowReport.replace(hour=0, minute=0, second=0, microsecond=0)

                    #calculate th f=differnce of days
                    diffSubmitdate = dateNowReport - submitDate
                    diffLastModified = dateNowReport - lastModifiedDate

                    #change th background of the cells with difference of days  
                    if diffSubmitdate.days >= 7:
                        sheetInc.cell(row, 2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid")                
                    if diffLastModified.days >= 3:
                        sheetInc.cell(row, 3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")

    
    #set the width of the columns
    sheetInc.column_dimensions[get_column_letter(1)].width = 25
    sheetInc.column_dimensions[get_column_letter(2)].width = 30
    sheetInc.column_dimensions[get_column_letter(3)].width = 30
    sheetInc.column_dimensions[get_column_letter(4)].width = 80
    sheetInc.column_dimensions[get_column_letter(5)].width = 17
    sheetInc.column_dimensions[get_column_letter(6)].width = 17
    sheetInc.column_dimensions[get_column_letter(7)].width = 30
    sheetInc.column_dimensions[get_column_letter(8)].width = 16
    sheetInc.column_dimensions[get_column_letter(10)].width = 16
    sheetInc.column_dimensions[get_column_letter(11)].width = 22
    sheetInc.column_dimensions[get_column_letter(12)].width = 18
    sheetInc.column_dimensions[get_column_letter(13)].width = 30
    sheetInc.column_dimensions[get_column_letter(14)].width = 30
    sheetInc.column_dimensions[get_column_letter(15)].width = 30
    sheetInc.column_dimensions[get_column_letter(16)].width = 30

    ######################### END CREATES INC SPREADSHEET ##############################


    ######################### START CREATES WO SPREADSHEET ###################################

        #Initialize header row, for the INC sheet Target Date
    for col in range(1, 12):
        #set the names for the header
        sheetWO.cell(1, col).value = sheetWOSource.cell(1, col).value 

        #fix the nam of column 1
        if col == 1:
            sheetWO.cell(1, col).value = "Work Order ID"
                
        #set the styes for the header
        sheetWO.cell(1, col).font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
        sheetWO.cell(1, col).border = copy(sheetWOSource.cell(1, col).border)
        sheetWO.cell(1, col).fill = copy(sheetWOSource.cell(1, col).fill)
        sheetWO.cell(1, col).number_format = copy(sheetWOSource.cell(1, col).number_format)
        sheetWO.cell(1, col).alignment = copy(sheetWOSource.cell(1, col).alignment)

    #loop for to check each rows and col of file
    for row in range(2, sheetWOSource.max_row + 1):

        # Edit - Naryan Sambhi 
        # - check if cell copied is in correct format, 
        # ie turn 45022.6358680556 serial number back into a date by copying existing formatting in source
        ###########################################################################################

        #cont loop for each col
        for col in range(1, 12):
            
            #set the value for each col according to source doc
            sheetWO.cell(row, col).value = sheetWOSource.cell(row, col).value

            #could be optomized a lot more than an inner for loop 
            for i in range (1, 12):
                sheetWO.cell(row, i).number_format = copy(sheetWOSource.cell(row, i).number_format)
                sheetWO.cell(row, i).alignment = copy(sheetWOSource.cell(row, i).alignment)



         ###########################################################################################


            #verify if the cell has hyperlink and set style to cell
            if sheetWOSource.cell(row, col).hyperlink:
                sheetWO.cell(row, col)._hyperlink = copy(sheetWOSource.cell(row, col).hyperlink)
                sheetWO.cell(row, col).style = incHyperlink

            #verify if the cell is even or not and set styles 
            if row % 2 == 0:
                sheetWO.cell(row, col).style = highlightEven
                if col == 1:
                    sheetWO.cell(row, col).style = incHyperlinkEven
            else:
                sheetWO.cell(row, col).style = highlightOdd

                if col == 1:
                    sheetWO.cell(row, col).style = incHyperlink
            


            #get the values of date to calculate difference of days
            submitDate = sheetWOSource.cell(row, 3).value
            lastModifiedDate = sheetWOSource.cell(row, 4).value
            
            #verify that the dates are not none
            if submitDate is not None and lastModifiedDate is not None:
                #remove hours of date to calculate dif
                submitDate = submitDate.replace(hour=0, minute=0, second=0, microsecond=0)
                lastModifiedDate = lastModifiedDate.replace(hour=0, minute=0, second=0, microsecond=0)
                dateNowReport = datetime.datetime.now()
                dateNowReport = dateNowReport.replace(hour=0, minute=0, second=0, microsecond=0)

                #get the dif of days
                dateNow = datetime.datetime.now()
                diffSubmitdate = dateNow - submitDate
                diffLastModified = dateNow - lastModifiedDate
                    
                #set the background of the cells according to difference of days
                if diffSubmitdate.days >= 7:
                    sheetWO.cell(row, 2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid")                
                if diffLastModified.days >= 3:
                    sheetWO.cell(row, 3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
                        # Check if the cell value is a serial number

                    
    #set the with of the columns
    sheetWO.column_dimensions[get_column_letter(1)].width = 25
    sheetWO.column_dimensions[get_column_letter(2)].width = 16
    sheetWO.column_dimensions[get_column_letter(3)].width = 25
    sheetWO.column_dimensions[get_column_letter(4)].width = 25
    sheetWO.column_dimensions[get_column_letter(5)].width = 25
    sheetWO.column_dimensions[get_column_letter(6)].width = 25
    sheetWO.column_dimensions[get_column_letter(7)].width = 80
    sheetWO.column_dimensions[get_column_letter(8)].width = 16
    sheetWO.column_dimensions[get_column_letter(10)].width = 16
    sheetWO.column_dimensions[get_column_letter(11)].width = 30

    



################################ END CREATES WO SPREADSHEET ########################################
        
############################################ END OF COPY PASTE FROM RECONREPORT.PY AND ONX V1 #################################

    if 'Sheet' in wbReport.sheetnames:  # remove default sheet
        wbReport.remove(wbReport['Sheet'])

    time.sleep(sleeptime)
    print("\n\nSaving file..")
    time.sleep(sleeptime)


    try:
        wbReport.save(save_path)
    except:
        print("\n************************************************\n\nERROR!\n\nCOULD NOT SAVE FILE, ENSURE DOCUMENT IS NOT OPEN, OR SAVE FILE PATH IS CORRECT\n\n************************************************")
        exit(1)

    print("\nFile Saved")
    time.sleep(sleeptime)


    




def createEmail(attachment): #attached file

    #create email
    print("\n\n*************************************************************************************************\n")

    print("Creating email...")

    time.sleep(3) #needed



    #current day ars
    dateNow = datetime.datetime.now()
    dateReport = dateNow.strftime("%b %d, %Y")
    dateReport2 = dateNow.strftime("%b %d, %Y")

    user_info = get_data(3)


    outlook = client.Dispatch("Outlook.Application")
    message = outlook.CreateItem(0)
    message.Display()

    message.To = messageTO
    message.CC = messageCC
    message.Subject = "Reconciliation Report ONX - " + dateReport

    message.Attachments.Add(attachment) #co-ordinate with previous code to check if file exists before email 


    html_body =  """
        <div >
            <p style="margin-bottom: 16px; font-size: 18px;" >Good Morning ONX team,</p>
            <p style="margin-bottom: 16px; font-size: 18px;">Please see the attached reconciliation report for today, """ + dateReport2 + """.<br></p>
            <p><span  style="background-color: red; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">Red for Priority/Important Tickects.</span><br>
            <span  style="background-color: green; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">Green for Resolved Tickets on ONX side.</span><br>
            <span  style="background-color: #FFC000; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">The tickets with orange highlights on the “submit date” column have been open for the past 7+ days.</span><br>
            <span  style="background-color: #FFFF00; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">The tickets with yellow highlights on the “last modified date” column have not been modified for the past 3+ days.</span><br>
            <span  style="background-color: blue; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">The tickets with Blue have more than 2 Hours difference in last modified date.</span><br>
            <span  style="background-color: purple; font-size: 18px; margin-bottom: 1px; padding-bottom: 1px">Purple for additional tickets in ONX queue that are not in eSMT.</span></p>
            <p style="margin-top: 32px'; margin-bottom: 60px; font-size: 18px;">Have a great day!!!</p>

       
    </div>

      <div style="margin-top: 50px; margin-bottom: 0; padding-bottom: 0;">

            <p style="color: blue; font-size: 14px; margin-left: 16px;  margin-bottom: 0;">
            """ + user_info + """ <br>
            Operations Control Centre – OCC <br>
            Ministry of Government and Consumer Services <br>
            Guelph Data Centre - GDC</p>

        </div>
         <div style="margin-left: 16px">
            <img  src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAEIAAAA9CAIAAACWQiirAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAABcBSURBVGhDtVoJdBzVle3qWru6elVLsmVJtuRF3hcMZjcGghPMZodtICQESMiEkDkhk3U4w8l+TpiQhIEEEyDgEMDEZglLPDbYBmQwOAa8ypYtyZLQLvXeXd219txfVZK6tZk5h7n6p9X1l/f+e//999/71dT3t738lw/36W6Xq1BwuVwU/iZBAX+lmKzzSE9K4K+eNffei9Y+8O7uv7UcK+Tzdv0UGEPTIjSWbwkoSqDc7p50Skc/0ykF05ysjPSxCzV55+JubvCg3W6KKq6cuGC2ZPQIHcMiZaKespomK3nDcHtZbngdXG63O69p0XQqmkmPFusxr2uYjC0/QFGU4XIl5Oz4zglZBj2aGu08NUAKkhYKBV3XZU2N5+RhaploLpvVVE3XDdMk3aCLiQBO1J0vbn7h6AEQA6GCYfgFT8QrQTOWaFYnitJNox+SaDrD0HYD6KJLhSQFBFHH92Fg9rKuDmTT4EzWQODXz55/78VrH2jctbn56Hijsicnq6qcyTCawguesrJI2COCjlEopNX8YCyaz2YUmuFFryR4IAhZomJACS4XM/ydMgwjl05+acmKX1yxAeao6lA3Ac8waSX/i+2vbnzvnUh5BRGXopI5eZ4/8MjVN6ysnaXout0TPDia2dPRevfWZzvTyXAoNKyKiUEU5HKlk3Emr6yonXXJnIazamctnVFbLkmYD8wqqeSbero+6upobGvZ39oclbP+YJB3MzA4h8QwHDEcFAoSxwZ4gXznrRoLfl5YNqPWZeiKpggMBz3BcGvDkZXVtWgq7gkEBZFxQ9ixnMaAdtOyqsiJ2Oxw5N/WbbjhjLMDHpGjaQjg9HC5YBdo/cKCJTDpt5qb/mv39r1tLWYg4BE82EJOJwslFgzlKYazCGOAGVcHw1lZdtM0jJgyjUUzqv0e0WkuAhZHw86bxI5twN4UQ8/FhtbNadh6x7e+eeGlFT4/lr1YhhGwNO3jhauWnrH19m99/dwLjVRSURVMw2m2MH4jTsx+ejC8uGaWqagQFfOM0OwZldOdtjGwXebk9oS5qoaR7u9bv2Dp72+8FVZU7DwAzTCgC32cQiHqb6/90h3nXqgkEqqqWpvCwXgxJgYcWq3P79I1bG6wCYpSfaTCaSPGOPUuKAJ6ut2JWHRFVfUP110zu7zSqbcAP/n2iWO/3v7Kj1/Zcv+O13c1N6VKvYLAsv++9koYQiqTMkxjZPWmEkPV9a54LKsq+B4QhDNrZzIeMaMouqLMiJQ3VFahHrstms3Ec1lrxOlBM0xGlrEDvrbmc2fNrHdqLbRHB+/a/OQ1j//3L3a89vDbb/78jdfWP/Hw155+tH1o0OlhYVYosn7FKsEowBE7VVOLkVaUnS3Nh3t78B1eZeXM2ZFQWMmk3S5q+bQZEdGL+pgsv9ve0h6LWSNOD/j/fCZ9cd2cyxoWOVUWUoryo5ee3/T2zhzNipFpUqTCUxbJe4QX9jb+9LUXulLJnKbJVsnp2mXz5leGQ3lDxwlsDy/1VKVQTeOjvp6Yopwzsw6Pfp6v8Yh9OVkMhqoj5Xaf9njs5aYj1y1Zbj9ODRyv2ZwMw7p44dL6MoeCjVc+2rfr8EdMJBIIBnEQwu6NguFl+ExV9daWY73PPO7jBM0kuwUrCWPSIC07aslTrQaDnWfoPfEh+7Fc8p1bO8ul5isl38oZNXblgJxtjg0h3rAfTwOwVzWJ4eaVRYo9CbzWSyePDTKM1+tzGcSfw+Y1JZ9PRGlVyWcyu/ftfaVx57b33kZ5vXHX9vcb80qe59gRIlOthsAwoqEf6jyVNwuCmwoInsXVM6HSaaK0sHya3adrsB/xR1icwPNOABKFGRImW+qphzKZvmzaxXIIwOxDOi7La2bNXj2znjhGbGXi/BzdsxQNu9p67PAnybjA8SRaKxSmEgOHEUszx3u7T8WjC8oiqJnu86OuKhCs8PvxCEs90t5qppI+wWONOD0Q7wg8uJccmXCvJLpxFSztkjlrcvaC2rp7LroMc8D0x5wn2CeNXR3N/b0Cx9k1UxkVDBT8Mjm5Y2jArqnx+c+oqZ0dLrMf4/lcc3QAYR3s1a45LSiaxuGlqarzbIGnadoNCjaRAkIAKBhzxcGHQrTpdhcXzupNTlhb8KnFQD+W5eKZ9J6W43aFVxAbZtZXhsP2Y3c89vEnnTzHMaVn6qSAYmh3Rs6kyUYfRZkklXu8LlVBXA4jITuDojwcC3NC3AmPhOUqPpqwdM7jsPamFMNSMwL0w9Ehe1jI612/dOWqGsffdyZig9FBGqH+p1yNQoFl2LSqtcSGRidFNiF7zeyGkKFns1kXQhLYs1dq7Dj1u7ff/M2ubXdt3vTXD9+Xdc3pPQIyJYfMlGKgE+12MVzrQF+LZVdh0XvNomVnW/4X+mke6ENk4hYE6MfqfxognhOxuQuFd44f7YyXHDVXrzxn9YIlanQggTTOLPh4/t22k/e99uK9Tz36/CtbYYUcwzpdJ8JUYhATRQdkiAN9ezraUAM9wQvbO6E7EX2/+Qg5F3nPaePZEdAuipd8O1pPvHXymFNloUwU77/25mvPvYiVZXmoLxUdVLMZxFo18xb85q7vfeWs89iSBbf2fFHFlFucdIcglFwwW2NRp3YY/enMkcFBF1ES5C3mMRUQVsLh6ob+2DtvHuzqdGotzCuf9sTNtz/31W/cc+nlt5235ruXrnvsxq/svvv7d66+RLJzh8kxtVFZYFlE74fa2zJF7gXr1BYb6lLyLo8IH/pphQCgR9MMhCJ729t+vf3VzniJdpC9XLV05c+u+OKDG278+br1X151fl04Mt4NEv06Xx2cRgyys3FCM0xTb9fR/l6nlmThuT0nmnQ54+KFArKL/wtAk2dZqbxyy4F/fnfLM839fSOhkQ2kkNj0SD/I0WYBLgvcBzKpIn9VIkiJGLAgrij05xiGhwwUhdMjmZdbBkbFSKnKsagTeGIEy4w6XJ5mLP1NJZth6B6O48vKXz7y8b888dCmD/YgQUX87zQXARLmNHXjOztv+dODx3t77XMQEyMsijiUroabJDSIzFEyqpLIyTAnBAgMzSZVbc+pVkTpdmt7dOhEdAibG1ThxJEVoDKhkM8sbI+CHku0NR7Iirwc56+cdjgeu/v5TV/88x9/teO1bU0Hu1OJlJpHFo6Y9+TQwNP/fO+25568d+uz2EidqbjNHTLrBRNMHFpYgNGbEbhD0ywXvSQ9steOovpy2a50inHTUJWHZuaHywQasaerV850JBNwgjieJJarD4Q8tBsumHVTKU1rTSaQq+BMPO3NCMSFj8PMVDnLahrn9VWXRaZ5RJZyK6bZmUkODA7gOOckMOFqJH+VKEJByN1akom0piChxyOWYlQMiObGJHK5vIwcCE8kn+Z5j18UYZHonEOsmU1bDfBPXEDygQa+44hNZ9IFQyfjTROnoV/ywS+Ty6XTiQGAM4hAg3Abed3MqzmXpttc3BznFQSGQoCPHIdKZTOakrPG0OAOJZIrEkuMkV1Edp6hGyLLhUPhcChEPoMhyUPuINCEyBRBByrRgk9QQeSDiaKepd0hfzAcKkP/UKgMMoCkFep9KpBdS8IFN9ZcYpkyrw88gDBIeiRsd5Kpo49h+D0i4WIxwlIXX/O4ZU0lR7VTaDf8A7msdAoUMdJqKWW0vjA8hNSTOqsJn4ivhoegkBtMyIp5FFWOL5CjAFsHF5sOIeWmGBy+Tj1hNNKEZS8aayIE+9oLz73ZehxL4+yHzxQwyivq5n73gov/8P6eF082FRSS1n/GIEZZoI729fZmU8T0nOrPDKCv6Np0X2BOWTkylq5UnKfZIsf/2cAkwTpFvdHSfCo2xDPwMZ+xIKCOLTk7VLaqeuaH3Z1w0BLPf+Zi4GQESeqWv23a3dZKRPh/MCoXz11ZN/d7F17y8N7GrSebXEpJtvTZgGwH+DSaRV6SM4wpSt4uJopZWkj9mM7FBX1wnsI5wCkjoR/TWlwmoY/icB/Tv6ToGnJCmrwKmbyQW2XiawwTcanmAF9MFJJCkzcEFHldMnagXSADNgmcy5j6kUJSVkKf3EaCLIpqFcDQdMLCgGt1uk1YyLkxhbGCPVwb6KfzuVgmFU8mM4l4LhHLJ2LZeCyRTMRSyQRiIeRl6Ao44z4trDFuHEuyko9l07F0Mg168ZiSiKOAUToRAwuwTisyZMFsJmMyeoqP2RtQIiqySk5Np70cX1VWPqssUh9CNCLCoadVFcFIa3SwJzqYyKQNnvd5fUj2QaJYL1Od4kQEKpvLK5mUx02Xh0JVofCsYGSGz+fjeAS3WU0blOUWuLhYtDc2KKsqI/nImxpEsMVZmnWKT3zBAxvI6Tp0H+S4q5edeePKs8+ZWe9HXEAOHhJNYq441/Ka1jzY9/rRw38/9OGJ/t6Cxxvw+cj7lclX2AbIaAUzMzTkY9jLFixdv3zlBXWzp/l8tBshARgQleNsByHEbIlcdn9nx4sHP/zHsUPR2FAQR7jbeek1gglWAzzyqiLHhi6on/8fl191Qd1ccfg6aEJgE/elEhsbd23c81bKZSKIAA+bzYSrAcuAhSTiQ8srp/903fqLGxZ5yEXbaUwS4fq+U60/enXrvq52f7gCCnXWxI6prD6jIDJoWj4Wu3nFWc/d/q9rGxZNLQPA0XRtqOwn6zY8dN1N4YIrHo+RSIGEdhPDpKh4InpmVc0fbrz1yiVnwGKLZYBS4NbwiQPBqbIAUS+at+CpL3/9/Fmzk8k4bKF4n9Arb7zu2GAfWQ1LMAxOx6LXLln2xy/dUe712Z1s4EjuTiT2dbQ1DfTBs4A9CcWdRmKHi6tqwl7vruamrGGKgoe4FwaxfWR13ey9naeOIM3SdYRK6UxGoOlfXnXd5QsWO4Mt5HXtjaOHnt6/9/XmY2+3njjQ0YZtUBUIQrNOD+s12qxI+RtHD8bJHa519WhJM7o38ADHlkylGiKR+664NiiUXLO2DvY/8MbrO9pbB/I56Fni+FWRim+cd9HnFy8nztQClHrtynPePdW66YN3DY9nYq/idiv53Krq2hVVM5yaYTz1fuN/PvdkTNUKkg/+hdZ0ryA+dtNXN6w4C6QcK6Vca+Y0XLJgyTP79+qCwDLOrXrRFkc6oqm8od101vmLpo/yQD9s39uf/OO+thN6uEIQeMyuPy+/emBfU8vx+2/5+oblZzpdXa6Q4LnhjLP/cfQQHGUoONGbWBh0oRDySt5xlx0+XvCIXpeawIwNmjEkjzo0eP///H1IyQU93pxK3tdBHLBIGTpDbo2tPNbS1agYcAuybswMRa5ZvMypspDMyb/fue291pP+mjqYEUwFY7Ea+gyhJTr08PvvIPrnKCpDXsa5IqI0kJd5SXKlsdsnkII4cppuj0f7U8maoHOJauP6FavqQ5EdJ5uOx2PNfT3dsaFsecXB3p57nnjYYDjT63UzLO0q0AUXK4pByWf5EUcORwysGjlJ87mlc+cvnF5tV9oAP4SPlN/Ps4wtA4AvDM2EKir39nRf9dhDZNGtehgDRTMMy4Z8fpjoeKtCaib6fEeGBl89emhZVQ3LjOqRY5hz58w7s262ZhrJbBaixhWlqa+n8fjhpuhQF47dTMolirzgBethGRyM7h7UcqZRHwziFHOqLBzt7/0knWQ9IpFheCT+QWxyBcgyqpvKUy6sBUrORSkkmSPU7J5jANlETmA4/oHd2x9p3Ak36jQMg6VpZKDTg6Fz6+asm7/oO6svffb2u/d858dbbv3Gnas/V+WR1HQSWhzjoEtmDE81RgbMJSpnEeGM7OMSQBKKAleRZUcKvJDTOjEoLEjA51dp+sevvXjbM0/ubmlOKQo87MjBVQycD16Oq/T6rly84sEv3rz1tm9+vmFhOhGHWyt2IaWTLpjFPwAB0DHsEUnyXloPoIly06qmp5PxbCKeTpGSxdJnM6CPVNPpVwqMwmRxKkASU/I/f3D/hscfuvzRB+97/cUXDu7/qLszJmcxRc0wxssksCyW6Nlbv3n5wuWZZEI3Rl8oFzlcWAU2XyqJNMS6L3OwsLKqSvJ1ZNOU1+uCysg0MBvyl1Zy0yX/2mVn8CyrWL8x8bJsfyb9ZtvJVD7nLX2lZAP7u2AauWRSh+ehCXecZXv7uvcf2s9UTPNK/jmStKRiek24rKGy8syaWVWBMF+6vCFR/MElaz/u7epX8gFBsCUZFqOAxJ1y88LBTzqO9/Uumk7eeduo8PnX1M95qnG36gtxDDJ/ssuh7JySzw8NfmHhst9efT0J16zOIPJ224n3ThwblLMI48ZoFAwz+byfF65esqJcEBC2WZFEwUMzTYMDb3V1yIhzk/H97W3kOsFFlfv96xevuOfitdXBcLEJQby68sruUycLvEAu3Ube/YEf+okM0018yIFiMcJe8XuXXXGo/dSB7g5XpBwGRtYN6cpA36KyyHUrzhJK3zwc6YFlRHmf36JaAkTaipzFpO84+4LV9XPhMex6CL/vk47exx9sSqd8lTOI9zDNnKGdSiZ+9/xffKb5w2uuxw60OwMQEgq16A9TsP/ZwKGYp92b//neycF+p8qyn0XTq/98x103rDo/TMPVmbRpiq7C5xav+P0td66ZO9/pZwFecufRgzBLSfTCszm1I6Ao+KieZPyTRIxnGA/L2gXfEeHevmZtJc/n+3uyiaicShSyGU7XqurqF82q95RqajCbGUin4NlB0K4Z3RtwkTTl9vsCR/p7f/X6S4/cckexmpfNqP3TLXe0Rwd70ynMLuIR6yMVYeuHCiMAhb/u27u9tVkKheycYNQOLJCfnYneoVx2y/73V9fPmzP8GwEAPb99yeXnzZ7/1smmtngsZxoSwzSEyy9duGTBtKpiiwJePPBhR18PwjYsI5iibVQMAKxhM55AaPPH+ySP52dXXY/95LRZwcKSqpolztNYwLds/Xjfb3duo0BB8BgaSQmdtmHAAhiK8gVCb5xoeuit7T9Ztz4kSnYTgEj53Po5Z86ss7wlubZhaHKV6DRbwKQ/7up45J2dGdMMsqNaHnsa4Cj38DwTCG58750vb3r03daTiNudtkkAEx/IpO9/8x93b3k67SoEg2FT18fLYAP5vIfjGH9g47tvfXvLX4/19ahFv2ABcPxZlsbBvRbLAAGyqrL5ow9uemrj8UQMyRO5QR7eXSWBuoNCgUUMzHJNXZ2vHDnQFh0IST4f8Z5YWzLSBnnVq2u9yQT8/Q9e/tvzH32gIAYJhYq3xPhAHTQwVuAE000f6Dj16pGPe5IJ+DSOtV/uEY9n04d2kFRgkTOK0pVK7Dh2+L5tL/1hz67evILMbFQGy+Amz8URZRVc6bxckDM+Xqwriyyurm2omBZAloOIwzS6U8nDPZ3H+3oGkimVcok+v8Aw9iQcEiAySS5uHzvwBKlsBoG7XxArAsG5ldPnlVdUeiUeztDKn9Kq0pmIn+jr/iQeG8qkZF0XJD95D0gRZ2aTwkRhUZOKARA5KbdhaKmcbGg6HBR2EglmSF/TMF06SQtokfeIgkCGF0tgYeoXAxZ9Cocx6JuaCvUSN2hdCKHVWhGXQRV0xEEUzfO8JAjklCnVlC2Ge8R5j4dFyEC+FvT6wn5EDwHW53d7Jbco0V6/GAgEA6GwLyDC3kjXyQlNAjLCLCCFRPoRDgSD/oDoD9JeH+ijMJKP8/ulQIi8dfD7YdUQjgSnE/FxY9+4cNpbipmwYBDGI5NkWIbnOGG4sCxcC7k5QgdCeNxAp6DVYj623irwR2gi9BGf0QjwR+mDF/kZB0PebxAuFqMxw+0C83Ivr5rhwaqBzeTFAfYKLHKkEnG7Pb8pC7kPxD8i6+jYMWUURfSJ5gkLewFGO48pEC7s8fwvJXnMrF6z6FQAAAAASUVORK5CYII=" style="width: 79px; margin-top: 0;"></p>
            
        </div>
   
    """

    message.HTMLBody = html_body  
    #message.Save()

    print("Email Has been created successfully")
    
    print("\n\n*************************************************************************************************\n")



###################################################################################################################################################
        


#def final report():

print("\nStarting..")

getFiles(2, subject1, UserEmail, UserFolder)

mergeFiles(ReconReportFP, ReconWorkOrderFP, ReconReportONX) #will take two files and merge accordingly 

createEmail(ReconReportONX) #will take file to attach, (maybe html body), (maybe to, cc, subject)

print("\nEnding..\n")